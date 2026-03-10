from fastapi import APIRouter, Depends, HTTPException, Path, UploadFile, File
from sqlalchemy.orm import Session
import openpyxl
import os
import pdfplumber
import shutil
import tempfile
from typing import List

from .. import models, schemas
from ..database import SessionLocal

router = APIRouter(prefix="/api/inventory", tags=["Inventario"])

EXCEL_FILE = "inventario_ml.xlsx"

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


@router.get("/excel", response_model=List[dict])
def get_excel_inventory():
    if not os.path.exists(EXCEL_FILE):
        return []
    
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        sheet = wb.active
        
        products = []
        headers = [cell.value for cell in sheet[1]]
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Si hay nombre de producto
                product = {
                    "Producto": row[0],
                    "Stock_Actual": row[1] if len(row) > 1 else 0,
                    "Lote_Inicial": row[2] if len(row) > 2 else 0,
                    "Precio": row[3] if len(row) > 3 else 0,
                    "Respuesta": calculate_availability(row[1] if len(row) > 1 else 0, row[2] if len(row) > 2 else 0)
                }
                products.append(product)
                
        return products
    except Exception as e:
        print(f"Error reading Excel: {e}")
        return []

def calculate_availability(stock, initial):
    if not initial or initial == 0:
        return "Consultar"
    try:
        percentage = (float(stock) / float(initial)) * 100
        if percentage > 75:
            return "Plena disponibilidad"
        elif percentage >= 25:
            return "Quedan pocas unidades"
        else:
            return "¡Últimas unidades!"
    except:
        return "Consultar"

@router.post("/import-pdf")
async def import_pdf_inventory(file: UploadFile = File(...)):
    # Save temp file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        shutil.copyfileobj(file.file, tmp)
        tmp_path = tmp.name

    extracted_data = []
    try:
        with pdfplumber.open(tmp_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        for row in table:
                            # Simple heuristic: filter out empty rows
                            if row and any(row):
                                extracted_data.append(row)
                if extracted_data:
                    pass
                else:
                    # If no tables, try to extract text line by line
                    text = page.extract_text()
                    if text:
                        lines = text.split('\n')
                        for line in lines:
                            extracted_data.append([line])
        
        # Simple heuristic mapping:
        if extracted_data and os.path.exists(EXCEL_FILE):
            wb = openpyxl.load_workbook(EXCEL_FILE)
            sheet = wb.active
            
            added_count = 0
            for row in extracted_data:
                product_name = None
                price = 0
                
                # Try to find Product and Price in the row
                for cell in row:
                    if not cell: continue
                    s_cell = str(cell).strip()
                    
                    # Detect Price (contains $ or is a reasonable number)
                    if '$' in s_cell or (s_cell.replace('.', '').replace(',', '').isdigit() and len(s_cell) < 10):
                        price = s_cell # Keep original formatting for now
                    
                    # Detect Product Name (longest string that is not a number/price)
                    elif len(s_cell) > 3 and not s_cell.replace('.','').isdigit() and '$' not in s_cell:
                        if product_name is None or len(s_cell) > len(product_name):
                            product_name = s_cell
                
                if product_name:
                    # Append to Excel: Producto, Stock_Actual, Lote_Inicial, Precio
                    # Note: Excel columns are 1-indexed. row[0]=Producto, row[1]=Stock, row[2]=Lote, row[3]=Precio
                    sheet.append([product_name, 0, 0, price])
                    added_count += 1
            
            wb.save(EXCEL_FILE)
            return {"data": extracted_data, "message": f"PDF procesado. {added_count} productos agregados al Excel."}

    except Exception as e:
        return {"error": str(e)}
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
    
    return {"data": extracted_data, "message": "PDF procesado, pero no se pudo guardar en Excel."}

@router.get("/products", response_model=list[schemas.LocalProduct])
def list_products(db: Session = Depends(get_db)):
    products = db.query(models.LocalProduct).order_by(models.LocalProduct.created_at.desc()).all()
    return products


@router.post("/demo-seed", response_model=list[schemas.LocalProduct])
def seed_demo_products(db: Session = Depends(get_db)):
    existing = db.query(models.LocalProduct).count()
    if existing > 0:
        return db.query(models.LocalProduct).all()

    demo_products = [
        models.LocalProduct(
            name="Auriculares Bluetooth Deportivos",
            sku="AURI-BT-001",
            category="Audio",
            price=54999,
            stock=25,
            image_url="https://via.placeholder.com/120",
            description="Auriculares inalámbricos con cancelación de ruido y estuche de carga rápida.",
        ),
        models.LocalProduct(
            name="Teclado Mecánico RGB Gamer",
            sku="TECL-RGB-002",
            category="Periféricos",
            price=79999,
            stock=10,
            image_url="https://via.placeholder.com/120",
            description="Teclado mecánico con switches lineales y retroiluminación RGB personalizable.",
        ),
        models.LocalProduct(
            name="Mochila Notebook 15.6'' Antirrobo",
            sku="MOCH-NB-003",
            category="Accesorios",
            price=45999,
            stock=18,
            image_url="https://via.placeholder.com/120",
            description="Mochila resistente al agua con puerto USB y compartimentos ocultos.",
        ),
    ]
    for p in demo_products:
        db.add(p)
    db.commit()
    for p in demo_products:
        db.refresh(p)
    return demo_products


@router.patch("/products/{product_id}/image", response_model=schemas.LocalProduct)
def update_product_image(
    product_id: int = Path(...),
    image_url: str | None = None,
    db: Session = Depends(get_db),
):
    product = db.query(models.LocalProduct).filter(models.LocalProduct.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    product.image_url = image_url
    db.add(product)
    db.commit()
    db.refresh(product)
    return product
